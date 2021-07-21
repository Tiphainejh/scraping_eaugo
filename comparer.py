import time
import os
import numpy as np
from sys import exit
from bs4 import BeautifulSoup
from selenium import webdriver
from datetime import datetime
import urllib3
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.workbook import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pandas as pd
import chromedriver_autoinstaller
chromedriver_autoinstaller.install()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def create_browser():
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    options.add_argument('log-level=3')

    return webdriver.Chrome(options=options)

products = []
productsPrices = []
comparison = []
df = pd.read_excel('vendeurs.xlsx')
sellers = df['Vendeurs'].tolist()

sellers.append("Eau-Go")
header = ["Produit","Référence"]
header.extend(sellers)
header.append("Ecart")

eaugo_price_id = 0
nStores = len(sellers) + 1
suivi = pd.read_excel("suivi.xlsx")
#code_shopping = pd.read_excel("code_shopping.xlsx") #ici

#fichier_produits = suivi.merge(code_shopping, on="gtin", how="outer") #ici

def compare_prices(product, eaugo_price_id):
    product_list = [p for p in product if isinstance(p, float)]
    eaugo_price = product[eaugo_price_id]
    
    # si les vendeurs ne sont pas presents sur la page
    if product_list == [] or eaugo_price == " ":
        product[-1] = 0
        product[-2] = -1
        return product

    min_price = min(product_list)
    min_price_id = product.index(min_price)

    # si eaugo est le moins cher
    if eaugo_price == min_price :
        prices = product[:-3]
        product_list = [p for p in prices if isinstance(p, float)]
        
        # s'il n'y a que eau-go dans la liste
        if product_list == []:
            product[-1] = 0
            product[-2] = -1
            return product
        sec_min_price = min(product_list) if prices != [] else min_price
        
        # si eau-go est le moins cher (vérif s'il y a deux fois le même prix minimum)
        if min_price_id == eaugo_price_id :
            product[-1] = (1 - eaugo_price / sec_min_price) * 100
            product[-2] = eaugo_price_id
        else:
            product[-1] = 0
            product[-2] = min_price_id
    # un autre vendeur est le moins cher
    else :
        product[-1] = (1 - min_price / eaugo_price) * 100
        product[-2] = min_price_id

    return product

def getPrices(products, priceRows, driver, startProduct = 0):
    for i, product in enumerate(products[startProduct:]) :
        i += startProduct
        try :
            googleShoppingCode = suivi["google_shopping"][i].replace("\"","")
        except :
            priceRows.append([product[0],product[1]])
            continue
        comp_page = "https://www.google.com/shopping/product/" + googleShoppingCode + "/offers"
        print(comp_page)
        buttonXPath = '//*[@id="yDmH0d"]/c-wiz/div/div/div[2]/div[1]/div[4]/form/div[1]/div/button'
        tableClass = "sh-osd__offer-row"
        priceID = "g9WBQb"
        searchID = "search"
        
        driver.get(comp_page)
        
        """
        try : 
            driver.get(comp_page)
        except TimeoutException as ex:
            priceRows = getPrices(products, priceRows, driver, i)
            return priceRows
        """
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        try :
            button = driver.find_element_by_xpath(buttonXPath)
            button.click()
        except:
            pass
        driver.get(comp_page)
        try :
            table = wait(driver, 20).until(EC.visibility_of_element_located((By.CLASS_NAME,tableClass)))

        except TimeoutException as ex:
            priceRows = getPrices(products, priceRows, driver, i)
            return priceRows

        table = driver.find_elements_by_class_name(tableClass)
        found = False
        priceRows.append([product[0],product[1]])
        for seller in sellers :
            for row in table :
                if seller in row.text:
                    price = row.find_element_by_class_name(priceID)
                    price = float(price.text.replace("\u20ac","").replace("\xa0","").replace("\u202f","").replace(",","."))
                    priceRows[i].append(price)
                    found = True
                    break
            if found == False:
                priceRows[i].append(" ")
            found = False

        
        priceRows[i].append(" ")
        priceRows[i].append(" ")
        # append sellers
        nSellers = len(sellers)

        time.sleep((30-5)*np.random.random()+5)
        print(str(i+1)+"/"+str(len(products)))
    return priceRows
    
def create_file(productPrices):
    global nStores 
    wb = load_workbook(filename = 'suivi.xlsx')
    sheet = wb.active
    rows = sheet.rows

    # rempli le fichier excel avec les vendeurs
    for r in range(2, sheet.max_row):
        products.append([sheet.cell(row=r, column=1).value])
        products[r-2].append(sheet.cell(row=r, column=2).value)
        products[r-2].append(sheet.cell(row=r, column=3).value)

    priceRows = []
    driver = create_browser()
    priceRows = getPrices(products, priceRows, driver)
    date = datetime.now().strftime("%Y-%m-%d")

    comparison_file = "comparaison.xlsx"
    if os.path.exists(comparison_file) :
        wb = load_workbook(comparison_file)
    else :
        wb = Workbook()

    if date in wb.sheetnames :
        date = datetime.now().strftime("%Y-%m-%d %H %M %S")

    wb.create_sheet(index = 0, title = date)
    ws = wb.worksheets[0]
    eaugo_price_id = header.index("Eau-Go")


    for j in range(1, len(priceRows)+1) :
        comparison = compare_prices(priceRows[j-1], eaugo_price_id)
        ws.cell(row=j, column=1).value = comparison[0]
        ws.cell(row=j, column=2).value = comparison[1]
        for s in range(3, 3+nStores):
            ws.cell(row=j, column=s).value = comparison[s-1]
            if s == (comparison[-2]+1):
                # si un autre vendeur est moins cher
                if comparison[-2] != (eaugo_price_id):
                    ws.cell(row=j, column=s).font = Font(bold=True, color = "FF0000")
                # si eaugo est moins cher
                else:
                    ws.cell(row=j, column=s).font = Font(bold=True, color = "F1C40F")
        # colonne des poucentages de prix
        ws.cell(row=j, column=3+nStores-1).value = str(int(comparison[-1])) + "%"

    ws.insert_rows(1)
    # mise en forme du fichier excel (header, taille des colonnes)
    for i, h in enumerate(header):
        ws.cell(row=1, column=i+1).value = h
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 10
    wb.save(comparison_file)
    return "Fin"


print(create_file(productsPrices))
exit(0)
#?prds=scoring:p
#?prds=epd